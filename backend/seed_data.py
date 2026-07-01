import csv
import os

from openpyxl import load_workbook
from database import default_password_for_user, get_connection, hash_password, init_db


BASE_DIR = os.path.dirname(os.path.dirname(__file__))
SAMPLE_DATA_DIR = os.path.join(BASE_DIR, "sample_data")
DATASET_XLSX = os.path.join(SAMPLE_DATA_DIR, "dataset.xlsx")
DEFAULT_CREDENTIALS_CSV = os.path.join(SAMPLE_DATA_DIR, "default_login_credentials.csv")
LD_SHARED_EMAIL = "scale@levelshift.com"
LD_SHARED_PASSWORD = "ld@321"


DEMO_USERS = [
    ("EMP001", "Arjun Sharma", "arjun.sharma@company.com"),
    ("EMP002", "Priya Nair", "priya.nair@company.com"),
    ("EMP003", "Ravi Kumar", "ravi.kumar@company.com"),
    ("EMP004", "Sneha Reddy", "sneha.reddy@company.com"),
    ("EMP005", "Kiran Mehta", "kiran.mehta@company.com"),
    ("EMP006", "Ananya Iyer", "ananya.iyer@company.com"),
    ("EMP007", "Deepak Joshi", "deepak.joshi@company.com"),
    ("EMP008", "Meera Pillai", "meera.pillai@company.com"),
    ("EMP009", "Suresh Patel", "suresh.patel@company.com"),
    ("EMP010", "Lakshmi Venkat", "lakshmi.venkat@company.com"),
    ("MGR001", "Rajesh Kapoor", "rajesh.kapoor@company.com"),
    ("MGR002", "Sunita Agarwal", "sunita.agarwal@company.com"),
    ("MGR003", "Vikram Singh", "vikram.singh@company.com"),
    ("LD001", "Kavitha Murthy", "kavitha.murthy@company.com"),
]

DEMO_REGISTRATION_IDS = ["REG001", "REG002", "REG003"]
DEMO_NOMINATION_IDS = ["NOM001", "NOM002"]

TRAININGS = [
    ("TRN001", "Advanced Python for Data Science", "Technical", "Online", "3 Days", "Dr. Anil Gupta", 20, "Python,Data Science,ML,Pandas", "Active", "Batch A: May 10-12 | Batch B: May 17-19 | Batch C: Jun 5-7"),
    ("TRN002", "Leadership & People Management", "Soft Skills", "Offline", "2 Days", "Prof. Meena Sharma", 15, "Leadership,Communication,Team Management", "Active", "Batch A: May 15-16 | Batch B: Jun 12-13"),
    ("TRN003", "Cloud Architecture on AWS", "Technical", "Hybrid", "4 Days", "Sudhir Rao", 12, "AWS,Cloud,DevOps,Architecture", "Active", "Batch A: May 20-23 | Batch B: Jun 17-20"),
    ("TRN004", "Power BI & Advanced Analytics", "Analytics", "Online", "2 Days", "Pooja Verma", 25, "Power BI,DAX,Data Visualization,Analytics", "Active", "Batch A: May 8-9 | Batch B: May 22-23 | Batch C: Jun 5-6"),
    ("TRN005", "Agile & Scrum Certification Prep", "Project Management", "Online", "3 Days", "Rohit Sharma", 30, "Agile,Scrum,JIRA,Sprint Planning", "Active", "Batch A: May 13-15 | Batch B: Jun 3-5"),
    ("TRN006", "Cybersecurity Fundamentals", "Security", "Online", "2 Days", "Amit Bose", 20, "Cybersecurity,Network Security,OWASP,Compliance", "Active", "Batch A: May 27-28 | Batch B: Jun 24-25"),
]


def load_from_excel(path):
    """Read dataset.xlsx Managers and Employees sheets into a flat list of DB-ready dicts.

    Supports two Excel formats:
      v1 (6 cols): Emp ID, Name, Domain, Role, Email ID, Password
      v2 (8 cols): Emp ID, Name, Domain, System Role, Email, Password, Designation, Business Unit (Managers)
                   Emp ID, Name, Domain, System Role, Email, Password, Designation, Manager ID  (Employees)
    Roles are derived from sheet name + domain when System Role column is absent.
    Manager ID is derived from domain when the Manager ID column is absent.
    """
    if not os.path.exists(path):
        print(f"WARNING: dataset.xlsx not found at {path}")
        return []

    # Copy to a temp file so we can read even if Excel has the file open
    import shutil, tempfile
    tmp_path = tempfile.mktemp(suffix=".xlsx")
    shutil.copy2(path, tmp_path)

    # Load without read_only so sheets can be scanned multiple times
    wb = load_workbook(tmp_path, read_only=False, data_only=True)

    def read_sheet_rows(ws):
        """Return (headers, list_of_value_tuples) for a sheet.
        Detects header row: if row 1 looks like a header (strings, no @) use row 1;
        otherwise skip rows 1-2 (old title/subtitle format) and use row 3."""
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not all_rows:
            return [], []
        r1 = all_rows[0]
        # Row 1 is a header if it has string-like values and no email addresses
        is_header = r1[0] and isinstance(r1[0], str) and "@" not in str(r1[0])
        if is_header:
            headers = [str(c).strip() if c else "" for c in r1]
            return headers, all_rows[1:]
        # Legacy: row 3 is header
        if len(all_rows) >= 3:
            headers = [str(c).strip() if c else "" for c in all_rows[2]]
            return headers, all_rows[3:]
        return [], []

    # ── First pass: build domain → manager_id from Managers sheet ──
    domain_to_mgr: dict = {}
    if "Managers" in wb.sheetnames:
        headers, data_rows = read_sheet_rows(wb["Managers"])
        for row in data_rows:
            if not row[0]:
                continue
            raw = dict(zip(headers, row))
            emp_id = str(raw.get("Emp ID") or "").strip()
            domain = str(raw.get("Domain") or "").strip()
            if emp_id and domain:
                domain_to_mgr[domain] = emp_id

    # ── Second pass: read new sheet format (L&D Team / Functional Head / Reporting Manager / Employees) ──
    entries = []
    name_to_id: dict = {}

    # Map new sheet names → system role
    NEW_SHEETS = {
        "Functional Head": "functional_head",
        "Reporting Manager": "manager",   # stored as manager in DB, maps to reporting_manager on login
        "L&D Team": "ld_admin",          # stored as ld_admin in DB, maps to ld_team on login
        "Employees": "employee",
    }

    for sheet_name, default_db_role in NEW_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        headers, data_rows = read_sheet_rows(wb[sheet_name])
        for row in data_rows:
            if not row[0]:
                continue
            raw = dict(zip(headers, row))

            # Normalise column names across sheets
            # Priority: sheet-specific ID column first, then generic
            if sheet_name == "Reporting Manager":
                emp_id = str(raw.get("RM ID") or raw.get("Emp ID") or "").strip()
            elif sheet_name == "Functional Head":
                emp_id = str(raw.get("FH ID") or raw.get("Emp ID") or "").strip()
            else:
                emp_id = str(raw.get("Emp ID") or "").strip()
            name   = str(raw.get("Name") or "").strip()
            domain = str(raw.get("Domain") or raw.get("Department") or "").strip()
            email  = str(raw.get("Email ID") or raw.get("Email") or "").strip()
            desig  = str(raw.get("Designation") or "").strip()
            mgr_id = str(raw.get("RM ID") or raw.get("Manager ID") or "").strip()
            # For employees, RM ID column is the manager; for RM the FH ID is manager
            if sheet_name == "Employees":
                mgr_id = str(raw.get("RM ID") or "").strip() or mgr_id
            elif sheet_name == "Reporting Manager":
                mgr_id = str(raw.get("FH ID") or "").strip()
            elif sheet_name == "Functional Head":
                mgr_id = ""

            # Override system role from sheet column if present
            # Map frontend role names to DB role names
            ROLE_TO_DB = {"reporting_manager": "manager", "ld_team": "ld_admin"}
            sys_role_col = str(raw.get("System Role") or "").strip().lower()
            raw_role = sys_role_col if sys_role_col else default_db_role
            db_role = ROLE_TO_DB.get(raw_role, raw_role)

            if not emp_id or not email:
                continue

            if name and emp_id:
                name_to_id[name] = emp_id

            entries.append({
                "employee_id": emp_id,
                "name": name,
                "email": email,
                "department": domain or "Learning & Development" if sheet_name == "L&D Team" else domain,
                "business_unit": domain or "Learning & Development",
                "designation": desig,
                "manager_id": mgr_id,
                "current_skills": "",
                "role": db_role,
            })

    # Legacy fallback: old sheet names (Managers / Employees) if new sheets absent
    if not any(s in wb.sheetnames for s in NEW_SHEETS):
        for sheet_name in ("Managers", "Employees"):
            if sheet_name not in wb.sheetnames:
                continue
            headers, data_rows = read_sheet_rows(wb[sheet_name])
            for row in data_rows:
                if not row[0]:
                    continue
                raw = dict(zip(headers, row))
                emp_id = str(raw.get("Emp ID") or "").strip()
                name   = str(raw.get("Name") or "").strip()
                domain = str(raw.get("Domain") or "").strip()
                email  = str(raw.get("Email") or raw.get("Email ID") or "").strip()
                desig  = str(raw.get("Designation") or "").strip()
                bunit  = str(raw.get("Business Unit") or domain).strip()
                sys_role = str(raw.get("System Role") or "").strip().lower()
                if not sys_role:
                    if sheet_name == "Managers":
                        sys_role = "ld_admin" if domain == "Learning & Development" else "manager"
                    else:
                        sys_role = "ld_admin" if domain == "Learning & Development" else "employee"
                mgr_id = str(raw.get("Manager ID") or "").strip()
                if not mgr_id and sheet_name == "Employees":
                    mgr_id = domain_to_mgr.get(domain, "")
                if not emp_id or not email:
                    continue
                entries.append({"employee_id":emp_id,"name":name,"email":email,"department":domain,
                    "business_unit":bunit,"designation":desig,"manager_id":mgr_id,"current_skills":"","role":sys_role})

    wb.close()
    try:
        os.remove(tmp_path)
    except OSError:
        pass
    return entries


def _placeholders(items):
    return ",".join("?" for _ in items)


def default_password_for_entry(entry):
    if (entry.get("role") or "") == "ld_admin":
        return LD_SHARED_PASSWORD
    return default_password_for_user(entry["employee_id"])


def write_default_credentials(entries):
    fieldnames = ["employee_id", "name", "email", "role", "default_password"]
    try:
        with open(DEFAULT_CREDENTIALS_CSV, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for entry in sorted(entries, key=lambda item: item["employee_id"]):
                writer.writerow(
                    {
                        "employee_id": entry["employee_id"],
                        "name": entry["name"],
                        "email": entry["email"],
                        "role": entry["role"],
                        "default_password": default_password_for_entry(entry),
                    }
                )
    except PermissionError:
        print(f"Skipping credential export because the file is locked: {DEFAULT_CREDENTIALS_CSV}")


def purge_demo_identity_data():
    init_db()
    conn = get_connection()
    c = conn.cursor()
    demo_user_ids = [employee_id for employee_id, _, _ in DEMO_USERS]
    demo_user_names = [name for _, name, _ in DEMO_USERS]

    existing_demo_users = []
    existing_demo_names = []
    for employee_id, name, email in DEMO_USERS:
        record = c.execute(
            "SELECT employee_id FROM employees WHERE employee_id=? AND lower(email)=lower(?) AND name=?",
            (employee_id, email, name),
        ).fetchone()
        if record:
            existing_demo_users.append(employee_id)
            existing_demo_names.append(name)

    demo_id_placeholders = _placeholders(demo_user_ids)
    reg_placeholders = _placeholders(DEMO_REGISTRATION_IDS)
    nom_placeholders = _placeholders(DEMO_NOMINATION_IDS)
    name_placeholders = _placeholders(demo_user_names)

    c.execute(
        f"DELETE FROM nomination_participants WHERE employee_id IN ({demo_id_placeholders}) OR nomination_id IN ({nom_placeholders})",
        demo_user_ids + DEMO_NOMINATION_IDS,
    )
    c.execute(
        f"DELETE FROM registration_requests WHERE employee_id IN ({demo_id_placeholders}) OR request_id IN ({reg_placeholders})",
        demo_user_ids + DEMO_REGISTRATION_IDS,
    )
    c.execute(
        f"DELETE FROM nomination_requests WHERE manager_id IN ({demo_id_placeholders}) OR nomination_id IN ({nom_placeholders})",
        demo_user_ids + DEMO_NOMINATION_IDS,
    )
    c.execute(
        f"DELETE FROM course_requests WHERE manager_id IN ({demo_id_placeholders})",
        demo_user_ids,
    )
    c.execute(
        f"DELETE FROM audit_logs WHERE entity_id IN ({reg_placeholders}) OR entity_id IN ({nom_placeholders}) OR performed_by IN ({name_placeholders})",
        DEMO_REGISTRATION_IDS + DEMO_NOMINATION_IDS + demo_user_names,
    )
    c.execute(
        f"UPDATE employees SET manager_id='' WHERE manager_id IN ({demo_id_placeholders})",
        demo_user_ids,
    )
    c.execute(
        f"UPDATE registration_requests SET reporting_manager='' WHERE reporting_manager IN ({name_placeholders})",
        demo_user_names,
    )

    if existing_demo_users:
        user_placeholders = _placeholders(existing_demo_users)
        c.execute(
            f"DELETE FROM employees WHERE employee_id IN ({user_placeholders})",
            existing_demo_users,
        )

    conn.commit()
    conn.close()


def seed():
    init_db()
    purge_demo_identity_data()
    conn = get_connection()
    c = conn.cursor()

    if c.execute("SELECT COUNT(*) FROM trainings").fetchone()[0] == 0:
        c.executemany(
            """INSERT INTO trainings
               (training_id, course_name, category, mode, duration, trainer_name,
                seats_available, skill_tags, status, batches)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            TRAININGS
        )

    roster_entries = load_from_excel(DATASET_XLSX)
    credential_entries = []
    roster_ids = {entry["employee_id"] for entry in roster_entries if entry.get("employee_id")}
    if roster_ids:
        existing_ids = [row[0] for row in c.execute("SELECT employee_id FROM employees").fetchall()]
        stale_ids = [employee_id for employee_id in existing_ids if employee_id not in roster_ids]
        if stale_ids:
            placeholders = _placeholders(stale_ids)
            c.execute(f"DELETE FROM employees WHERE employee_id IN ({placeholders})", stale_ids)
    for entry in roster_entries:
        if not entry["employee_id"] or not entry["email"]:
            continue
        default_password_hash = hash_password(default_password_for_entry(entry))
        credential_entries.append(entry)
        existing = c.execute("SELECT employee_id FROM employees WHERE employee_id=?", (entry["employee_id"],)).fetchone()
        if existing:
            c.execute(
                """UPDATE employees
                SET name=?, email=?, department=?, business_unit=?, designation=?, manager_id=?, current_skills=?, role=?, password_hash=?
                WHERE employee_id=?""",
                (
                    entry["name"],
                    entry["email"],
                    entry["department"],
                    entry["business_unit"],
                    entry["designation"],
                    entry["manager_id"],
                    entry["current_skills"],
                    entry["role"],
                    default_password_hash,
                    entry["employee_id"],
                ),
            )
            continue

        c.execute(
            """INSERT INTO employees
            (employee_id,name,email,department,business_unit,designation,manager_id,current_skills,role,password_hash)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                entry["employee_id"],
                entry["name"],
                entry["email"],
                entry["department"],
                entry["business_unit"],
                entry["designation"],
                entry["manager_id"],
                entry["current_skills"],
                entry["role"],
                default_password_hash,
            ),
        )

    conn.commit()
    conn.close()
    write_default_credentials(credential_entries)
    print("Reference training data loaded successfully.")


if __name__ == "__main__":
    seed()
