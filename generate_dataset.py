"""
generate_dataset.py - LevelShift LMS Dataset Generator (v2)
# -*- coding: utf-8 -*-
Sheets: L&D Team | Functional Head | Reporting Manager | Employees
Run:   python generate_dataset.py
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv, os

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
SAMPLE_DIR = os.path.join(BASE_DIR, "sample_data")
os.makedirs(SAMPLE_DIR, exist_ok=True)

def pwd(emp_id): return f"LS@{emp_id}#2026"
def email(name):
    p = name.lower().replace("'","").replace("-","").split()
    return f"{p[0]}.{p[-1]}@levelshift.com"

LD_SHARED_EMAIL = "scale@levelshift.com"
LD_SHARED_PASSWORD = "ld@321"
LD_MANAGER_REPORTS_TO = ("FH003", "Sunita Krishnan")

# -- 1. L&D TEAM (1 manager + 20 members) --------------------------------------
LD_MANAGER = ("LD001", "Kavitha Murthy", "Head of Learning & Development")
LD_MEMBERS = [
    ("LD002","Bhavya Joseph","L&D Program Manager"),
    ("LD003","Charan Deep","LMS Administrator"),
    ("LD004","Divya Mathew","Learning Experience Designer"),
    ("LD005","Ekansh Suri","Training Operations Analyst"),
    ("LD006","Fiona Dsilva","Assessment Specialist"),
    ("LD007","Girish Prakash","Instructional Designer"),
    ("LD008","Hema Ramanathan","Content Strategist"),
    ("LD009","Ishita Sharma","E-Learning Developer"),
    ("LD010","Jayant Pillai","Training Delivery Specialist"),
    ("LD011","Kavitha Rajput","Curriculum Developer"),
    ("LD012","Latha Subramaniam","Learning Analytics Specialist"),
    ("LD013","Mahesh Deekshit","Knowledge Management Analyst"),
    ("LD014","Nalini Krishnamurthy","Facilitation Specialist"),
    ("LD015","Omkar Phadke","Learning Technology Analyst"),
    ("LD016","Preethi Reddy","Talent Development Partner"),
    ("LD017","Qasim Siddiqui","Vendor & Content Specialist"),
    ("LD018","Ranjith Kumar","Coaching & Mentoring Lead"),
    ("LD019","Shalini Menon","Learning Evaluation Analyst"),
    ("LD020","Tara Iyer","Digital Learning Specialist"),
    ("LD021","Uma Chandrasekhar","Programme Management Lead"),
]

# -- 2. FUNCTIONAL HEADS (5, each owning 4 domains) ----------------------------
FUNCTIONAL_HEADS = [
    ("FH001","Arjun Nambiar","VP - Technology",["Software Engineering","Cloud & DevOps","Quality Engineering","Cybersecurity"]),
    ("FH002","Deepika Sethi","VP - Data & Analytics",["Data Science & AI","Business Intelligence","Data Engineering","ML Platform"]),
    ("FH003","Ramesh Pillai","VP - Business Operations",["Finance & Compliance","HR & People Ops","Legal & Risk","Procurement"]),
    ("FH004","Sunita Krishnan","VP - Product & Growth",["Product Management","Sales & Customer Success","Marketing & Brand","Customer Experience"]),
    ("FH005","Vikram Anand","VP - Infrastructure & Security",["Network & Systems","Information Security","IT Operations","Business Continuity"]),
]
DOMAIN_TO_FH = {}
for fh_id,fh_name,_,domains in FUNCTIONAL_HEADS:
    for d in domains:
        DOMAIN_TO_FH[d] = (fh_id, fh_name)

# -- 3. REPORTING MANAGERS (20, one per domain) ---------------------------------
REPORTING_MANAGERS = [
    ("RM001","Aarthi Raman","Software Engineering","Engineering Manager"),
    ("RM002","Bharath Menon","Data Science & AI","Data Science Manager"),
    ("RM003","Charu Sinha","Cybersecurity","Cybersecurity Manager"),
    ("RM004","Deepak Arora","Cloud & DevOps","DevOps Manager"),
    ("RM005","Eesha Nair","Product Management","Product Manager"),
    ("RM006","Farhan Qureshi","Quality Engineering","QA Manager"),
    ("RM007","Gayathri Iyer","Finance & Compliance","Finance Manager"),
    ("RM008","Harish Babu","HR & People Ops","HR Manager"),
    ("RM009","Ishita Kapoor","Sales & Customer Success","Sales Manager"),
    ("RM010","Jayesh Rao","Business Intelligence","BI Manager"),
    ("RM011","Kavya Subramaniam","Data Engineering","Data Engineering Manager"),
    ("RM012","Lokesh Dutta","ML Platform","ML Platform Manager"),
    ("RM013","Meera Nair","Network & Systems","Network Manager"),
    ("RM014","Nikhil Wagle","Information Security","InfoSec Manager"),
    ("RM015","Onamika Roy","IT Operations","IT Ops Manager"),
    ("RM016","Pradeep Menon","Legal & Risk","Legal & Risk Manager"),
    ("RM017","Queenie Rodrigues","Procurement","Procurement Manager"),
    ("RM018","Rajiv Kumar","Marketing & Brand","Marketing Manager"),
    ("RM019","Shalini Bose","Customer Experience","CX Manager"),
    ("RM020","Tilak Sehgal","Business Continuity","BC Manager"),
]

# -- 4. EMPLOYEES (10 per domain = 200 total) -----------------------------------
EMPLOYEES_BY_DOMAIN = {
    "Software Engineering":[
        ("Aditi Kulkarni","Senior Software Engineer"),("Bhavesh Rao","Software Engineer"),
        ("Chaitra Sen","Backend Engineer"),("Darshan Pillai","Build Engineer"),("Ekta Joseph","Solutions Engineer"),
        ("Faizal Hussain","Frontend Developer"),("Gaurav Sharma","Full Stack Developer"),("Hema Krishnan","Platform Engineer"),
        ("Ishaan Mehta","API Developer"),("Jayesh Patel","Systems Engineer"),
    ],
    "Data Science & AI":[
        ("Akash Reddy","Data Scientist"),("Bhavana Mishra","ML Engineer"),("Chetan Deshpande","Data Analyst"),
        ("Divya Subramaniam","Analytics Engineer"),("Eshan Kapoor","Applied Scientist"),("Farah Shaikh","AI Research Analyst"),
        ("Ganesh Rao","Feature Engineer"),("Hina Qureshi","NLP Engineer"),("Irfan Siddiqui","Computer Vision Engineer"),
        ("Jyoti Sharma","BI Analyst"),
    ],
    "Cybersecurity":[
        ("Kavya Ramesh","Security Analyst"),("Lokesh Mehta","IAM Engineer"),("Megha Tandon","AppSec Engineer"),
        ("Nikhil Joshi","Threat Hunter"),("Onamika Singh","Security Compliance Analyst"),("Pradeep Nambiar","Penetration Tester"),
        ("Queenie Fernandes","SOC Analyst"),("Rajiv Suresh","Vulnerability Analyst"),("Shalini Bhat","Cloud Security Engineer"),
        ("Tilak Kapoor","Network Security Engineer"),
    ],
    "Cloud & DevOps":[
        ("Pranav Shetty","DevOps Engineer"),("Qurrat Mirza","Site Reliability Engineer"),("Ritika Salgaonkar","Cloud Analyst"),
        ("Satvik Anand","Platform Engineer"),("Tanvi Deshpande","Observability Engineer"),("Ujjwal Pathak","Infrastructure Engineer"),
        ("Vandana Rao","Cloud Architect"),("Waseem Shaikh","Kubernetes Engineer"),("Xavier Pinto","GitOps Engineer"),
        ("Yash Dubey","CI/CD Specialist"),
    ],
    "Product Management":[
        ("Uday Ghosh","Associate Product Manager"),("Vaishali Sood","Product Analyst"),("Wasim Ali","Product Operations Specialist"),
        ("Xenia Lazarus","Technical Product Owner"),("Yashica Bedi","Growth Product Associate"),("Zuber Kureshi","Product Strategist"),
        ("Alisha Prabhu","Product Designer"),("Biju Nair","Product Marketing Manager"),("Chandana Pillai","User Research Lead"),
        ("Dhanush Reddy","Product Metrics Analyst"),
    ],
    "Quality Engineering":[
        ("Zubin Contractor","Automation Test Engineer"),("Anirudh Pal","QA Engineer"),("Brinda Sethi","Performance Tester"),
        ("Cedric Dsouza","Mobile QA Analyst"),("Devika Bhardwaj","Test Lead"),("Fatima Sheikh","Regression Test Analyst"),
        ("Govind Nair","Test Automation Architect"),("Harpreet Kaur","API Quality Analyst"),("Ishaan Verma","Security Tester"),
        ("Jayashri Krishnan","Accessibility Tester"),
    ],
    "Finance & Compliance":[
        ("Falguni Mehra","Compliance Officer"),("Gokul Nair","Accounts Executive"),("Heena Parashar","Revenue Analyst"),
        ("Imran Noor","Internal Audit Associate"),("Jaya Venkataraman","Financial Analyst"),("Krishnan Pillai","Tax Analyst"),
        ("Leela Krishnamurthy","Treasury Analyst"),("Mahesh Kulkarni","Cost Accountant"),("Namrata Sontakke","Budget Planning Analyst"),
        ("Ojus Vaidya","Financial Controller"),
    ],
    "HR & People Ops":[
        ("Janani Krishnan","Talent Acquisition Specialist"),("Keshav Lulla","HR Generalist"),("Leena Purohit","L&D Coordinator"),
        ("Mihir Sachdeva","People Analyst"),("Navya Bhasin","Performance Management Analyst"),("Omer Ansari","Compensation & Benefits Analyst"),
        ("Pallavi Deshmukh","HR Business Partner"),("Qadir Shah","Workforce Planning Analyst"),("Rani Devi","Employee Engagement Specialist"),
        ("Surendra Mane","Talent Management Analyst"),
    ],
    "Sales & Customer Success":[
        ("Omkar Kulshreshtha","Account Executive"),("Parul Grover","Customer Success Manager"),("Quincy Dsilva","Renewals Specialist"),
        ("Raashi Talwar","Sales Development Representative"),("Sameer Bopanna","Growth Analyst"),("Tanushree Rao","Enterprise Sales Executive"),
        ("Umang Shah","Inside Sales Specialist"),("Veena Kumar","Partner Success Manager"),("Wilton Dcosta","Customer Onboarding Specialist"),
        ("Yasmin Merchant","Key Account Manager"),
    ],
    "Business Intelligence":[
        ("Aarav Sharma","BI Developer"),("Bhakti Joshi","Dashboard Analyst"),("Chetan Nair","Reporting Specialist"),
        ("Deepa Menon","Data Visualisation Analyst"),("Esha Pillai","BI Architect"),("Farida Shaikh","Analytics Engineer"),
        ("Girish Rao","SQL Developer"),("Heer Patel","Tableau Developer"),("Indira Kumar","Power BI Analyst"),
        ("Jai Mehta","Looker Developer"),
    ],
    "Data Engineering":[
        ("Karthik Iyer","Data Pipeline Engineer"),("Lavanya Nair","ETL Developer"),("Mohan Reddy","Data Platform Engineer"),
        ("Nandita Bose","Spark Engineer"),("Omkar Kale","Airflow Engineer"),("Pooja Srivastava","Data Warehouse Analyst"),
        ("Rajeev Saxena","Kafka Engineer"),("Snehal Dharap","dbt Engineer"),("Tarika Chawla","Data Quality Analyst"),
        ("Uma Shankar","Data Lake Architect"),
    ],
    "ML Platform":[
        ("Vivek Dubey","MLOps Engineer"),("Wasim Baig","Model Deployment Engineer"),("Xavier Lobo","Feature Store Engineer"),
        ("Yamini Bhat","Experiment Tracking Specialist"),("Zaid Ansari","Model Monitoring Analyst"),("Abhinav Wagh","GPU Infrastructure Engineer"),
        ("Balaji Iyer","Vector DB Engineer"),("Chandan Roy","LLM Fine-tuning Specialist"),("Deepika Jain","Inference Optimisation Engineer"),
        ("Eknath Desai","AI Platform Analyst"),
    ],
    "Network & Systems":[
        ("Feroze Irani","Network Engineer"),("Geeta Krishnamurthy","Systems Administrator"),("Hemang Trivedi","Network Architect"),
        ("Isha Bhatia","Network Security Analyst"),("Jigar Patel","Wireless Network Engineer"),("Kiran Pawar","Network Operations Analyst"),
        ("Laxman Reddy","DC Network Engineer"),("Meenal Pillai","SDN Engineer"),("Naveen Kumar","DNS & DHCP Specialist"),
        ("Oindrila Das","Network Automation Engineer"),
    ],
    "Information Security":[
        ("Poornima Shetty","InfoSec Analyst"),("Rahul Bendre","Vulnerability Manager"),("Savita Patil","GRC Analyst"),
        ("Tapan Ghosh","SIEM Analyst"),("Uma Kapoor","DLP Engineer"),("Vinayak Joshi","Identity & Access Manager"),
        ("Waqar Ahmad","Threat Intelligence Analyst"),("Yash Malhotra","Red Team Analyst"),("Zainab Kadri","Security Architect"),
        ("Abhinav Sreenivasan","Cyber Risk Analyst"),
    ],
    "IT Operations":[
        ("Bakhtavar Khan","IT Support Engineer"),("Chirag Singhvi","IT Service Manager"),("Diya Bhatt","Desktop Support Analyst"),
        ("Eshan Rathi","ITSM Analyst"),("Fizza Hussain","Asset Management Analyst"),("Girija Narayanan","IT Procurement Analyst"),
        ("Heer Vadia","End User Computing Analyst"),("Ishan Roy","IT Operations Analyst"),("Jasmine Wadia","Change Management Analyst"),
        ("Kartik Bhardwaj","Problem Management Analyst"),
    ],
    "Legal & Risk":[
        ("Lahari Reddy","Legal Counsel"),("Mrinal Bose","Contract Specialist"),("Nitin Kalia","Risk Analyst"),
        ("Ojas Mehta","Compliance Manager"),("Priya Singh","IP Counsel"),("Qasim Mirza","Regulatory Affairs Analyst"),
        ("Ravi Iyer","Corporate Secretary"),("Swapna Rajagopalan","Legal Operations Analyst"),("Tejinder Walia","Data Privacy Officer"),
        ("Urvashi Joshi","Contract Manager"),
    ],
    "Procurement":[
        ("Vimal Malhotra","Procurement Analyst"),("Wasim Kolapkar","Vendor Manager"),("Nirupama Thomas","Category Manager"),
        ("Yuvraj Mahajan","Sourcing Specialist"),("Zubair Patel","Contract Negotiator"),("Anupama Devi","Spend Analyst"),
        ("Baldev Singh","Purchase Order Analyst"),("Chinmay Phadke","Supplier Relations Manager"),("Daksha Mehta","Procurement Ops Analyst"),
        ("Ela Sharma","Cost Management Analyst"),
    ],
    "Marketing & Brand":[
        ("Falguni Shah","Brand Manager"),("Gokul Menon","Digital Marketing Analyst"),("Heena Kaur","Content Strategist"),
        ("Imran Siddiqui","SEO Specialist"),("Jaya Patel","Social Media Manager"),("Krishnan Rao","Performance Marketing Analyst"),
        ("Leela Sharma","Campaign Manager"),("Mahesh Joshi","Email Marketing Specialist"),("Namrata Iyer","Growth Hacker"),
        ("Ojus Kumar","PR Specialist"),
    ],
    "Customer Experience":[
        ("Priyamvada Nayak","CX Manager"),("Qasim Ansari","Customer Insights Analyst"),("Ravi Bose","Voice of Customer Analyst"),
        ("Swathi Pillai","CX Operations Specialist"),("Tarun Shah","NPS Analyst"),("Urvashi Menon","Customer Journey Analyst"),
        ("Varsha Rao","CX Automation Analyst"),("Wasim Gupta","Customer Feedback Manager"),("Xavier Sharma","CX Strategy Analyst"),
        ("Yamuna Krishnan","CX Program Manager"),
    ],
    "Business Continuity":[
        ("Zahira Siddiqui","BC Analyst"),("Alok Srivastava","Disaster Recovery Specialist"),("Bindu Menon","Crisis Management Analyst"),
        ("Chittaranjan Das","BC Plan Coordinator"),("Dilip Joshi","Risk & Resilience Analyst"),("Esha Pandey","BCP Tester"),
        ("Fariyal Baig","Incident Response Analyst"),("Gopal Shinde","BC Documentation Specialist"),("Harini Balakrishnan","Emergency Response Analyst"),
        ("Indumati Nair","Supply Chain Resilience Analyst"),
    ],
}

rm_by_domain = {rm[2]:(rm[0],rm[1]) for rm in REPORTING_MANAGERS}
emp_records = []
emp_num = 1
for domain, emps in EMPLOYEES_BY_DOMAIN.items():
    rm_id, rm_name = rm_by_domain.get(domain,("",""))
    fh_id, fh_name = DOMAIN_TO_FH.get(domain,("",""))
    for name, designation in emps:
        emp_id = f"EMP{emp_num:03d}"
        emp_records.append({"emp_id":emp_id,"name":name,"domain":domain,"designation":designation,
            "email":email(name),"password":pwd(emp_id),"manager_id":rm_id,"manager_name":rm_name,
            "fh_id":fh_id,"fh_name":fh_name})
        emp_num += 1

# -- Excel styling -------------------------------------------------------------
def make_header(ws, headers, fill_hex, row=1):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill=fill; c.font=font; c.alignment=align; c.border=border

def auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx+4, 45)

wb = Workbook()
wb.remove(wb.active)

# Sheet 1 ??? L&D Team
ws = wb.create_sheet("L&D Team")
make_header(ws, ["Emp ID","Name","Designation","Department","Email ID","Password","System Role","Reports To"], "1565C8")
mid,mname,mdesig = LD_MANAGER
ws.append([mid, mname, mdesig, "Learning & Development", LD_SHARED_EMAIL, LD_SHARED_PASSWORD, "ld_team", LD_MANAGER_REPORTS_TO[1]])
for eid,name,desig in LD_MEMBERS:
    ws.append([eid, name, desig, "Learning & Development", LD_SHARED_EMAIL, LD_SHARED_PASSWORD, "ld_team", mname])
auto_width(ws); ws.freeze_panes="A2"

# Sheet 2 � Functional Head
ws = wb.create_sheet("Functional Head")
make_header(ws, ["FH ID","Name","Designation","Email ID","Password","System Role","Domains Managed","Reporting Managers Under"], "6D28D9")
rm_by_fh = {}
for rm_id,rm_name,domain,_ in REPORTING_MANAGERS:
    fh_id,_ = DOMAIN_TO_FH.get(domain,("",""))
    rm_by_fh.setdefault(fh_id,[]).append(rm_name)
for fh_id,fh_name,fh_desig,domains in FUNCTIONAL_HEADS:
    ws.append([fh_id, fh_name, fh_desig, email(fh_name), pwd(fh_id), "functional_head",
               ", ".join(domains), ", ".join(rm_by_fh.get(fh_id,[]))])
auto_width(ws); ws.freeze_panes="A2"

# Sheet 3 � Reporting Manager
ws = wb.create_sheet("Reporting Manager")
make_header(ws, ["RM ID","Name","Domain","Designation","Email ID","Password","System Role","Reports To (FH)","FH ID"], "0F766E")
for rm_id,rm_name,domain,rm_desig in REPORTING_MANAGERS:
    fh_id,fh_name = DOMAIN_TO_FH.get(domain,("",""))
    ws.append([rm_id, rm_name, domain, rm_desig, email(rm_name), pwd(rm_id), "reporting_manager", fh_name, fh_id])
auto_width(ws); ws.freeze_panes="A2"

# Sheet 4 � Employees
ws = wb.create_sheet("Employees")
make_header(ws, ["Emp ID","Name","Domain","Designation","Email ID","Password","System Role",
                 "Reporting Manager","RM ID","Functional Head","FH ID"], "B45309")
for r in emp_records:
    ws.append([r["emp_id"],r["name"],r["domain"],r["designation"],r["email"],r["password"],"employee",
               r["manager_name"],r["manager_id"],r["fh_name"],r["fh_id"]])
auto_width(ws); ws.freeze_panes="A2"

out = os.path.join(SAMPLE_DIR,"dataset.xlsx")
wb.save(out)
print(f"Saved: {out}")
print(f"  L&D Team: {1+len(LD_MEMBERS)}  |  FHs: {len(FUNCTIONAL_HEADS)}  |  RMs: {len(REPORTING_MANAGERS)}  |  Employees: {len(emp_records)}")

# Backend CSVs
p = os.path.join(SAMPLE_DIR,"managers.csv")
with open(p,"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(["manager_id","manager_name","email","domain","designation","fh_id","fh_name"])
    for rm_id,rm_name,domain,rm_desig in REPORTING_MANAGERS:
        fh_id,fh_name = DOMAIN_TO_FH.get(domain,("",""))
        w.writerow([rm_id,rm_name,email(rm_name),domain,rm_desig,fh_id,fh_name])

p = os.path.join(SAMPLE_DIR,"functional_heads.csv")
with open(p,"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(["fh_id","fh_name","email","designation","domains"])
    for fh_id,fh_name,fh_desig,domains in FUNCTIONAL_HEADS:
        w.writerow([fh_id,fh_name,email(fh_name),fh_desig,"|".join(domains)])

p = os.path.join(SAMPLE_DIR,"employees.csv")
with open(p,"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(["employee_id","employee_name","email","domain","designation","manager_id","manager_name"])
    for r in emp_records:
        w.writerow([r["emp_id"],r["name"],r["email"],r["domain"],r["designation"],r["manager_id"],r["manager_name"]])

p = os.path.join(SAMPLE_DIR,"ld_team.csv")
with open(p,"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(["employee_id","name","email","department","designation","manager_id","role"])
    w.writerow([mid,mname,email(mname),"Learning & Development",mdesig,"","ld_team"])
    for eid,name,desig in LD_MEMBERS:
        w.writerow([eid,name,email(name),"Learning & Development",desig,mid,"ld_team"])

p = os.path.join(SAMPLE_DIR,"default_login_credentials.csv")
with open(p,"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(["Role","Emp ID","Name","Email","Password"])
    w.writerow(["L&D Team (Head)",mid,mname,LD_SHARED_EMAIL,LD_SHARED_PASSWORD])
    for eid,name,_ in LD_MEMBERS:
        w.writerow(["L&D Team",eid,name,LD_SHARED_EMAIL,LD_SHARED_PASSWORD])
    for fh_id,fh_name,_,_ in FUNCTIONAL_HEADS:
        w.writerow(["Functional Head",fh_id,fh_name,email(fh_name),pwd(fh_id)])
    for rm_id,rm_name,domain,_ in REPORTING_MANAGERS:
        w.writerow(["Reporting Manager",rm_id,rm_name,email(rm_name),pwd(rm_id)])
    for r in emp_records[:10]:
        w.writerow(["Employee",r["emp_id"],r["name"],r["email"],r["password"]])

print("Done! Run: cd backend && python seed_data.py")
